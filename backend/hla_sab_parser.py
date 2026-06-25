
import os
import re
import io
import zipfile
from datetime import datetime, date



def sab_kit_id(name: str) -> str:
    s = str(name or "").strip().lower()
    if "lambda" in s or "kit 2" in s or "kit2" in s:
        return "kit2"
    return "kit1"



_AUTO_PRA_RE = re.compile(r"^\s*The SAB % PRA Class (?:I|II) is \d+%\.?\s*$",
                           re.IGNORECASE)


def sab_pra_sentence(pct_text, sab_class) -> str:
    m = re.search(r"\d+", str(pct_text or ""))
    if not m:
        return ""
    cls = "II" if str(sab_class or "").strip().upper().endswith("II") else "I"
    return f"The SAB % PRA Class {cls} is {int(m.group())}%."


def is_auto_pra_text(text) -> bool:
    return bool(_AUTO_PRA_RE.match(str(text or "")))



def parse_sab_allele_text(text: str) -> list:
    result = []
    for line in (text or "").strip().splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(.*)[,\t]\s*([0-9]+(?:\.[0-9]+)?)\s*$", line)
        if not m:
            continue
        allele = m.group(1).strip().rstrip(",").strip()
        if not allele:
            continue
        try:
            mfi = int(float(m.group(2)))
        except ValueError:
            continue
        result.append((allele, mfi))
    return sorted(result, key=lambda x: -x[1])



def parse_sab_excel(path: str, kit: str = "kit1") -> dict:
    if sab_kit_id(kit) == "kit2":
        return _parse_sab_excel_kit2(path)
    return _parse_sab_excel_kit1(path)


def _parse_sab_excel_kit1(path: str) -> dict:
    def _fmt(v):
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            return v.strftime("%d-%m-%Y")
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    if os.path.splitext(path)[1].lower() == ".xls":
        import xlrd

        class _XlrdCell:
            def __init__(self, value):
                self.value = value

        class _XlrdSheet:
            def __init__(self, sh, datemode):
                self._sh = sh
                self._dm = datemode
                self.title = sh.name
                self.max_row = sh.nrows
                self.max_column = sh.ncols
                self._images = []

            def cell(self, row, col):
                r, c = row - 1, col - 1
                if r < 0 or r >= self._sh.nrows or c < 0 or c >= self._sh.ncols:
                    return _XlrdCell(None)
                ct = self._sh.cell_type(r, c)
                val = self._sh.cell_value(r, c)
                if ct == xlrd.XL_CELL_DATE:
                    t = xlrd.xldate_as_tuple(val, self._dm)
                    val = date(*t[:3]) if t[3:] == (0, 0, 0) else datetime(*t)
                elif ct == xlrd.XL_CELL_EMPTY:
                    val = None
                return _XlrdCell(val)

        xlrd_wb = xlrd.open_workbook(path)

        class _FakeWb:
            def __init__(self, sheets):
                self.worksheets = sheets

        wb = _FakeWb([_XlrdSheet(xlrd_wb.sheet_by_index(i), xlrd_wb.datemode)
                      for i in range(xlrd_wb.nsheets)])
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)

    pat_ws = rep_ws = chart_ws = None
    for ws in wb.worksheets:
        t = (ws.title or "").lower()
        if pat_ws is None and "patient" in t:
            pat_ws = ws
        elif rep_ws is None and "report" in t:
            rep_ws = ws
        elif chart_ws is None and ("chart" in t or "chat" in t or "bead" in t):
            chart_ws = ws
    if rep_ws is None:
        cand = [w for w in wb.worksheets if w not in (pat_ws, chart_ws)]
        rep_ws = max(cand, key=lambda w: w.max_row * w.max_column) if cand else None

    HEADER_MAP = {
        "patient name":           "patient_name",
        "gender/ age":            "gender_age",
        "gender / age":           "gender_age",
        "gender/age":             "gender_age",
        "hospital mr no":         "hospital_mr_no",
        "specimen":               "specimen",
        "hospital/clinic":        "hospital_clinic",
        "hospital / clinic":      "hospital_clinic",
        "pin":                    "pin",
        "sample number":          "sample_number",
        "sample collection date": "collection_date",
        "sample receipt date":    "receipt_date",
        "report date":            "report_date",
    }
    patient = {}
    if pat_ws is not None:
        hdr_row = None
        for r in range(1, min(pat_ws.max_row, 15) + 1):
            for c in range(1, pat_ws.max_column + 1):
                v = pat_ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() == "patient name":
                    hdr_row = r
                    break
            if hdr_row:
                break
        if hdr_row:
            for c in range(1, pat_ws.max_column + 1):
                hv = pat_ws.cell(hdr_row, c).value
                if not isinstance(hv, str):
                    continue
                key = HEADER_MAP.get(hv.strip().lower())
                if key:
                    patient[key] = _fmt(pat_ws.cell(hdr_row + 1, c).value)

    alleles, pra_pct, sab_class = [], None, None
    if rep_ws is not None:
        title = f" {(rep_ws.title or '').upper()} "
        if "SAB II" in title or "CLASS II" in title or " II " in title:
            sab_class = "II"
        elif "SAB I" in title or "CLASS I" in title or " I " in title:
            sab_class = "I"

        ant_col = raw_col = hdr_row = None
        for r in range(1, rep_ws.max_row + 1):
            labels = {}
            for c in range(1, rep_ws.max_column + 1):
                v = rep_ws.cell(r, c).value
                if isinstance(v, str):
                    labels[v.strip().lower()] = c
            if "antigens" in labels and "raw value" in labels:
                ant_col, raw_col, hdr_row = labels["antigens"], labels["raw value"], r
                break
        if hdr_row:
            for r in range(hdr_row + 1, rep_ws.max_row + 1):
                ant = rep_ws.cell(r, ant_col).value
                raw = rep_ws.cell(r, raw_col).value
                if raw is None:
                    continue
                allele = str(ant).strip() if ant is not None else ""
                if not allele:
                    continue
                try:
                    alleles.append((allele, int(round(float(raw)))))
                except (TypeError, ValueError):
                    continue
            alleles.sort(key=lambda x: -x[1])

        for r in range(1, rep_ws.max_row + 1):
            for c in range(1, rep_ws.max_column + 1):
                v = rep_ws.cell(r, c).value
                if isinstance(v, str) and "% pra" in v.strip().lower():
                    for c2 in range(c + 1, rep_ws.max_column + 1):
                        nv = rep_ws.cell(r, c2).value
                        if isinstance(nv, (int, float)):
                            pra_pct = float(nv)
                            break
                    break
            if pra_pct is not None:
                break

    chart_bytes, chart_rot = None, 0
    try:
        zf = zipfile.ZipFile(path)
        media = [n for n in zf.namelist() if n.startswith("xl/media/")]
        if media:
            chart_name = max(media, key=lambda n: zf.getinfo(n).file_size)
            chart_bytes = zf.read(chart_name)
            chart_base = os.path.basename(chart_name)
            for n in zf.namelist():
                if not re.match(r"xl/drawings/drawing\d+\.xml$", n):
                    continue
                rels = f"xl/drawings/_rels/{os.path.basename(n)}.rels"
                if rels not in zf.namelist():
                    continue
                rels_xml = zf.read(rels).decode("utf-8", "ignore")
                targets = re.findall(r'Target="([^"]+)"', rels_xml)
                if not any(os.path.basename(t) == chart_base for t in targets):
                    continue
                draw_xml = zf.read(n).decode("utf-8", "ignore")
                rots = re.findall(r'<a:xfrm[^>]*\brot="(-?\d+)"', draw_xml)
                if rots:
                    chart_rot = int(rots[0])
                break
        zf.close()
    except Exception:
        chart_bytes, chart_rot = chart_bytes, 0
    if chart_bytes is None:
        best = 0
        for ws in wb.worksheets:
            for im in getattr(ws, "_images", []):
                try:
                    blob = im._data()
                except Exception:
                    continue
                if blob and len(blob) > best:
                    best, chart_bytes = len(blob), blob

    if chart_bytes and chart_rot:
        try:
            from PIL import Image as PILImage
            deg = (chart_rot / 60000.0) % 360
            if deg:
                pi = PILImage.open(io.BytesIO(chart_bytes))
                pi = pi.rotate(-deg, expand=True)
                buf = io.BytesIO()
                pi.save(buf, format="PNG")
                chart_bytes = buf.getvalue()
        except Exception:
            pass

    return {
        "patient":     patient,
        "alleles":     alleles,
        "chart_bytes": chart_bytes,
        "pra_pct":     pra_pct,
        "sab_class":   sab_class,
    }


def _parse_sab_excel_kit2(path: str) -> dict:
    def _s(v):
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    def _num(v):
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = _s(v)
        return float(s) if re.fullmatch(r"-?\d+(?:\.\d+)?", s) else None

    rows = []
    if os.path.splitext(path)[1].lower() == ".xls":
        import xlrd
        sh = xlrd.open_workbook(path).sheet_by_index(0)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]
    else:
        import openpyxl
        sh = openpyxl.load_workbook(path, data_only=True).worksheets[0]
        rows = [list(r) for r in sh.iter_rows(values_only=True)]

    allele_col = raw_col = hdr = None
    for ri, row in enumerate(rows):
        labels = {}
        for ci, v in enumerate(row):
            s = _s(v).lower()
            if s and s not in labels:
                labels[s] = ci
        if "allele equiv" in labels and "raw" in labels:
            allele_col, raw_col, hdr = labels["allele equiv"], labels["raw"], ri
            break

    alleles = []
    if hdr is not None:
        for row in rows[hdr + 1:]:
            allele = _s(row[allele_col]) if allele_col < len(row) else ""
            mfi = _num(row[raw_col]) if raw_col < len(row) else None
            if not allele or mfi is None:
                continue
            alleles.append((allele, int(round(mfi))))
        alleles.sort(key=lambda x: -x[1])

    blob = " ".join(_s(v).upper() for row in rows for v in row if _s(v))
    if "LS2A" in blob or "SAB II" in blob:
        sab_class = "II"
    elif "LS1A" in blob or "SAB I" in blob:
        sab_class = "I"
    else:
        sab_class = None

    pra_pct = None
    for row in rows:
        hit = next((ci for ci, v in enumerate(row) if "%pra" in _s(v).lower()),
                   None)
        if hit is None:
            continue
        for v2 in row[hit + 1:]:
            n = _num(v2)
            if n is not None:
                pra_pct = n
                break
        break

    patient = {}
    for row in rows:
        hit = next((ci for ci, v in enumerate(row)
                    if _s(v).lower().startswith("sample id")), None)
        if hit is None:
            continue
        val = next((_s(v2) for v2 in row[hit + 1:] if _s(v2)), "")
        if val:
            patient["pin"] = val
        break

    return {
        "patient":     patient,
        "alleles":     alleles,
        "chart_bytes": None,
        "pra_pct":     pra_pct,
        "sab_class":   sab_class,
    }
